from __future__ import annotations

from datetime import date, datetime
from html import escape
from io import BytesIO
from typing import Any

from fastapi_modulo.modulos.control_interno.repositorios.base import session_scope
from fastapi_modulo.modulos.control_interno.repositorios.reporte_repository import (
    list_acciones,
    list_actividades,
    list_controles,
    list_evidencias,
    list_hallazgos,
)
from fastapi_modulo.modulos.control_interno.servicios._common import contains_text, iso_date

REPORT_SECTIONS = {
    "controles": "Catalogo de controles",
    "actividades": "Programa anual - Actividades",
    "evidencias": "Evidencias de control",
    "hallazgos": "Hallazgos",
    "acciones": "Acciones correctivas",
}
LOGO_URL = "/modulos/control_interno/static/description/control.svg"


def _in_range(value, desde: str | None, hasta: str | None) -> bool:
    if not value:
        return not desde and not hasta
    current = value if isinstance(value, date) else date.fromisoformat(str(value)[:10])
    if desde and current < date.fromisoformat(desde):
        return False
    if hasta and current > date.fromisoformat(hasta):
        return False
    return True


def _tipo_label(tipo: str) -> str:
    return {
        "completo": "Informe completo",
        "controles": "Catalogo de controles",
        "programa": "Programa anual",
        "evidencias": "Evidencias",
        "hallazgos": "Hallazgos y acciones",
    }.get(tipo, tipo)


def _filters_text(filters: dict[str, Any]) -> str:
    labels = {
        "anio": "Ano",
        "componente_coso": "Componente COSO",
        "nivel_riesgo": "Nivel de riesgo",
        "estado_hallazgo": "Estado hallazgo",
        "estado_actividad": "Estado actividad",
        "estado_control": "Estado control",
        "resultado_ev": "Resultado evidencia",
        "control_id": "Control",
        "fecha_desde": "Desde",
        "fecha_hasta": "Hasta",
        "q": "Busqueda",
    }
    parts = [f"{labels[key]}: {value}" for key, value in filters.items() if value not in (None, "", []) and key in labels]
    return " | ".join(parts) or "Sin filtros aplicados"


def _section_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    return rows[:10]


def datos_reporte_service(
    *,
    tipo: str = "completo",
    anio: int | None = None,
    componente_coso: str | None = None,
    nivel_riesgo: str | None = None,
    estado_hallazgo: str | None = None,
    estado_actividad: str | None = None,
    estado_control: str | None = None,
    resultado_ev: str | None = None,
    control_id: int | None = None,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    q: str | None = None,
    version: str = "detallada",
) -> dict[str, Any]:
    incluir = {
        "controles": tipo in ("controles", "completo"),
        "programa": tipo in ("programa", "completo"),
        "evidencias": tipo in ("evidencias", "completo"),
        "hallazgos": tipo in ("hallazgos", "completo"),
        "acciones": tipo in ("hallazgos", "completo"),
    }
    filters = {
        "anio": anio,
        "componente_coso": componente_coso,
        "nivel_riesgo": nivel_riesgo,
        "estado_hallazgo": estado_hallazgo,
        "estado_actividad": estado_actividad,
        "estado_control": estado_control,
        "resultado_ev": resultado_ev,
        "control_id": control_id,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "q": q,
    }
    result: dict[str, Any] = {
        "tipo": tipo,
        "version": version,
        "titulo": _tipo_label(tipo),
        "filtros": filters,
        "filtros_texto": _filters_text(filters),
        "logo_url": LOGO_URL,
        "generado_en": datetime.now().strftime("%d/%m/%Y %H:%M"),
    }
    with session_scope() as db:
        if incluir["controles"]:
            rows = []
            for item in list_controles(db):
                if componente_coso and item.componente != componente_coso:
                    continue
                if estado_control and item.estado != estado_control:
                    continue
                if not contains_text(item.codigo or "", item.nombre or "", item.area or "", query=q):
                    continue
                rows.append({
                    "ID": item.id,
                    "Codigo": item.codigo,
                    "Nombre": item.nombre,
                    "Componente COSO": item.componente,
                    "Area": item.area,
                    "Tipo de riesgo": item.tipo_riesgo or "",
                    "Periodicidad": item.periodicidad,
                    "Estado": item.estado,
                    "Normativa": item.normativa or "",
                    "Creado": iso_date(item.creado_en),
                })
            result["controles"] = _section_summary(rows) if version == "resumida" else rows
        if incluir["programa"]:
            rows = []
            for item in list_actividades(db):
                programa_anio = item.programa.anio if item.programa else None
                if anio and programa_anio != anio:
                    continue
                if estado_actividad and item.estado != estado_actividad:
                    continue
                if control_id and item.control_id != control_id:
                    continue
                rows.append({
                    "ID": item.id,
                    "Programa": item.programa.nombre if item.programa else "",
                    "Ano": programa_anio or "",
                    "Control vinculado": item.control.codigo if item.control else "",
                    "Descripcion": item.descripcion or "",
                    "Responsable": item.responsable or "",
                    "F. inicio programada": iso_date(item.fecha_inicio_programada),
                    "F. fin programada": iso_date(item.fecha_fin_programada),
                    "F. inicio real": iso_date(item.fecha_inicio_real),
                    "F. fin real": iso_date(item.fecha_fin_real),
                    "Estado": item.estado,
                    "Observaciones": item.observaciones or "",
                })
            result["actividades"] = _section_summary(rows) if version == "resumida" else rows
        if incluir["evidencias"]:
            rows = []
            for item in list_evidencias(db):
                if resultado_ev and item.resultado_evaluacion != resultado_ev:
                    continue
                if control_id and item.control_id != control_id:
                    continue
                if not _in_range(item.fecha_evidencia, fecha_desde, fecha_hasta):
                    continue
                rows.append({
                    "ID": item.id,
                    "Titulo": item.titulo,
                    "Tipo": item.tipo,
                    "Fecha": iso_date(item.fecha_evidencia),
                    "Control vinculado": item.control.codigo if item.control else "",
                    "Resultado evaluacion": item.resultado_evaluacion,
                    "Observaciones": item.observaciones or "",
                    "Archivo": item.archivo_nombre_original or item.archivo_nombre or "",
                    "Creado": iso_date(item.creado_en),
                })
            result["evidencias"] = _section_summary(rows) if version == "resumida" else rows
        if incluir["hallazgos"]:
            rows = []
            for item in list_hallazgos(db):
                if nivel_riesgo and item.nivel_riesgo != nivel_riesgo:
                    continue
                if estado_hallazgo and item.estado != estado_hallazgo:
                    continue
                if componente_coso and item.componente_coso != componente_coso:
                    continue
                if control_id and item.control_id != control_id:
                    continue
                if not _in_range(item.fecha_deteccion, fecha_desde, fecha_hasta):
                    continue
                rows.append({
                    "ID": item.id,
                    "Codigo": item.codigo or "",
                    "Titulo": item.titulo,
                    "Componente COSO": item.componente_coso or "",
                    "Control vinculado": item.control.codigo if item.control else "",
                    "Nivel de riesgo": item.nivel_riesgo,
                    "Estado": item.estado,
                    "Causa": item.causa or "",
                    "Efecto": item.efecto or "",
                    "Fecha deteccion": iso_date(item.fecha_deteccion),
                    "Fecha limite": iso_date(item.fecha_limite),
                    "Responsable": item.responsable or "",
                    "Acciones": len(item.acciones or []),
                    "Creado": iso_date(item.creado_en),
                })
            result["hallazgos"] = _section_summary(rows) if version == "resumida" else rows
        if incluir["acciones"]:
            rows = []
            for item in list_acciones(db):
                if not _in_range(item.fecha_compromiso, fecha_desde, fecha_hasta):
                    continue
                rows.append({
                    "ID": item.id,
                    "Hallazgo codigo": item.hallazgo.codigo if item.hallazgo else "",
                    "Hallazgo titulo": item.hallazgo.titulo if item.hallazgo else "",
                    "Descripcion accion": item.descripcion,
                    "Responsable": item.responsable or "",
                    "Estado": item.estado,
                    "F. compromiso": iso_date(item.fecha_compromiso),
                    "F. ejecucion": iso_date(item.fecha_ejecucion),
                    "Evidencia seguimiento": item.evidencia_seguimiento or "",
                })
            result["acciones"] = _section_summary(rows) if version == "resumida" else rows
    result["resumen"] = {key: len(result.get(key, [])) for key in REPORT_SECTIONS}
    return result


def exportar_excel_service(data: dict[str, Any]) -> tuple[bytes, str, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=11, color="1E3A5F")
    alt_fill = PatternFill(fill_type="solid", fgColor="F0F4FA")
    thin_border = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"), top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def write_sheet(ws, title: str, rows: list[dict[str, Any]]):
        ws.append([title])
        ws.cell(1, 1).font = title_font
        ws.append([f"Generado: {data['generado_en']}"])
        ws.append([f"Filtros: {data['filtros_texto']}"])
        if not rows:
            ws.append(["Sin datos para los filtros aplicados."])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        header_row = ws.max_row
        for idx, col in enumerate(headers, 1):
            cell = ws.cell(header_row, idx, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
        for row_idx, row in enumerate(rows, ws.max_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col in enumerate(headers, 1):
                cell = ws.cell(row_idx, col_idx, row.get(col, ""))
                if fill:
                    cell.fill = fill
                cell.alignment = left
                cell.border = thin_border
        for col_idx, col in enumerate(headers, 1):
            width = max((len(str(row.get(col, ""))) for row in rows), default=0)
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(width, len(col)) + 3, 50)

    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    summary_sheet.append(["Reporte de Control Interno"])
    summary_sheet.append([data["titulo"]])
    summary_sheet.append([f"Version: {data['version']}"])
    summary_sheet.append([f"Generado: {data['generado_en']}"])
    summary_sheet.append([f"Filtros: {data['filtros_texto']}"])
    summary_sheet.append([])
    summary_sheet.append(["Seccion", "Registros"])
    for key, label in REPORT_SECTIONS.items():
        if key in data:
            summary_sheet.append([label, len(data.get(key, []))])

    for key, label in REPORT_SECTIONS.items():
        if key not in data:
            continue
        write_sheet(workbook.create_sheet(title=label[:31]), label, data.get(key, []))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"CI_Reporte_{data['tipo']}_{data['version']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename


def exportar_html_service(data: dict[str, Any], *, printable: bool = True) -> str:
    sections_html = ""
    for key, label in REPORT_SECTIONS.items():
        rows = data.get(key)
        if rows is None:
            continue
        if not rows:
            sections_html += f"<section class=\"seccion\"><h2>{escape(label)}</h2><p>Sin datos para los filtros aplicados.</p></section>"
            continue
        headers = list(rows[0].keys())
        header_html = "".join(f"<th>{escape(str(header))}</th>" for header in headers)
        body_html = ""
        for index, row in enumerate(rows):
            cells = "".join(f"<td>{escape(str(row.get(header, '')))}</td>" for header in headers)
            body_html += f"<tr class=\"{'alt' if index % 2 else ''}\">{cells}</tr>"
        sections_html += f"""
        <section class="seccion">
          <h2>{escape(label)} <span>({len(rows)} registros)</span></h2>
          <div class="table-wrap">
            <table>
              <thead><tr>{header_html}</tr></thead>
              <tbody>{body_html}</tbody>
            </table>
          </div>
        </section>
        """
    print_button = '<button class="no-print" onclick="window.print()">Imprimir</button>' if printable else ""
    return f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{escape(data['titulo'])}</title>
  <style>
    body{{font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#222;margin:0;padding:24px 28px;background:#f5f7fb}}
    .sheet{{max-width:1200px;margin:0 auto;background:#fff;padding:24px 28px;border:1px solid #dbe3ef}}
    .header{{display:flex;align-items:center;justify-content:space-between;gap:20px;border-bottom:3px solid #1E3A5F;padding-bottom:16px;margin-bottom:16px}}
    .header img{{width:72px;height:72px;object-fit:contain}}
    .header h1{{margin:0;color:#1E3A5F;font-size:22px}}
    .meta{{color:#5a6678;font-size:10px}}
    .chips{{display:flex;gap:8px;flex-wrap:wrap;margin:14px 0 18px}}
    .chip{{background:#eef3fb;border:1px solid #d6e0f1;border-radius:999px;padding:6px 10px;font-size:10px;color:#24456e}}
    .summary{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin:16px 0 22px}}
    .card{{border:1px solid #d8e1ef;background:#f9fbfe;padding:12px 14px}}
    .card strong{{display:block;color:#1E3A5F;font-size:17px}}
    .card span{{font-size:10px;color:#59657a}}
    .seccion{{margin-bottom:24px}}
    .seccion h2{{background:#1E3A5F;color:#fff;padding:8px 12px;margin:0 0 8px;font-size:12px}}
    .seccion h2 span{{font-weight:400;opacity:.8}}
    table{{width:100%;border-collapse:collapse;font-size:10px}}
    thead tr{{background:#2d5494;color:#fff}}
    th,td{{border:1px solid #d8dee8;padding:5px 7px;text-align:left;vertical-align:top}}
    tbody tr.alt{{background:#f0f4fa}}
    .table-wrap{{overflow:auto}}
    button{{padding:8px 14px;background:#1E3A5F;color:#fff;border:none;border-radius:4px;cursor:pointer}}
    @media print{{body{{background:#fff;padding:0}}.sheet{{border:none;max-width:none}}.no-print{{display:none!important}}@page{{margin:1.5cm}}}}
  </style>
</head>
<body>
  <div class="sheet">
    <div class="header">
      <div style="display:flex;align-items:center;gap:16px;">
        <img src="{escape(data['logo_url'])}" alt="Logo">
        <div>
          <h1>{escape(data['titulo'])}</h1>
          <div class="meta">Sistema de Control Interno | Version {escape(data['version'])} | Generado: {escape(data['generado_en'])}</div>
        </div>
      </div>
      {print_button}
    </div>
    <div class="chips">
      <div class="chip">{escape(data['filtros_texto'])}</div>
    </div>
    <div class="summary">
      {"".join(f'<div class="card"><strong>{value}</strong><span>{escape(REPORT_SECTIONS.get(key, key))}</span></div>' for key, value in data["resumen"].items() if key in data)}
    </div>
    {sections_html or '<p>No hay datos para los filtros seleccionados.</p>'}
  </div>
</body>
</html>"""


def exportar_pdf_service(data: dict[str, Any]) -> tuple[bytes, str, str]:
    lines = [
        "Sistema de Control Interno",
        data["titulo"],
        f"Version: {data['version']}",
        f"Generado: {data['generado_en']}",
        f"Filtros: {data['filtros_texto']}",
        "",
    ]
    for key, label in REPORT_SECTIONS.items():
        if key not in data:
            continue
        rows = data.get(key, [])
        lines.append(f"{label}: {len(rows)} registros")
        for row in rows[:15]:
            preview = " | ".join(f"{column}: {value}" for column, value in list(row.items())[:4])
            lines.append(preview[:110])
        lines.append("")
    filename = f"CI_Reporte_{data['tipo']}_{data['version']}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return _build_simple_pdf(lines), "application/pdf", filename


def _build_simple_pdf(lines: list[str]) -> bytes:
    objects = []
    content = ["BT", "/F1 12 Tf", "50 790 Td"]
    for index, line in enumerate(lines):
        if index == 0:
            content.append(f"({_escape_pdf(line)}) Tj")
            continue
        content.append("0 -18 Td")
        content.append(f"({_escape_pdf(line)}) Tj")
    content.append("ET")
    stream = "\n".join(content).encode("latin-1", errors="replace")
    objects.append(b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n")
    objects.append(b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n")
    objects.append(b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n")
    objects.append(b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n")
    objects.append(f"5 0 obj << /Length {len(stream)} >> stream\n".encode("latin-1") + stream + b"\nendstream endobj\n")
    buffer = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(buffer))
        buffer.extend(obj)
    xref_offset = len(buffer)
    buffer.extend(f"xref\n0 {len(offsets)}\n".encode("latin-1"))
    buffer.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        buffer.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    buffer.extend((f"trailer << /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF").encode("latin-1"))
    return bytes(buffer)


def _escape_pdf(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


__all__ = [
    "datos_reporte_service",
    "exportar_excel_service",
    "exportar_html_service",
    "exportar_pdf_service",
]
