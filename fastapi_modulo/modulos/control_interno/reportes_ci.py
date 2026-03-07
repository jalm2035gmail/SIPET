from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Request, Response
from fastapi.responses import HTMLResponse, JSONResponse

router = APIRouter()

BASE_TEMPLATE_PATH = os.path.join("fastapi_modulo", "templates", "modulos", "control_interno")


def _bind_core_symbols() -> None:
    if globals().get("_CORE_BOUND"):
        return
    from fastapi_modulo import main as core
    globals()["render_backend_page"] = getattr(core, "render_backend_page")
    globals()["_CORE_BOUND"] = True


def _load_template(filename: str) -> str:
    path = os.path.join(BASE_TEMPLATE_PATH, filename)
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return fh.read()
    except OSError:
        return "<p>No se pudo cargar la vista.</p>"


# ── Página ────────────────────────────────────────────────────────────────────

@router.get("/control-interno/reportes", response_class=HTMLResponse)
def reportes_ci_page(request: Request):
    _bind_core_symbols()
    return render_backend_page(
        request,
        title="Reportes de control interno",
        description="Generación y exportación de reportes del sistema de control interno COSO",
        content=_load_template("reportes_ci.html"),
        hide_floating_actions=True,
        show_page_header=False,
    )


# ── API: datos (JSON preview) ─────────────────────────────────────────────────

@router.get("/api/ci-reporte/datos")
def api_datos(
    tipo:             Optional[str] = "completo",
    anio:             Optional[int] = None,
    componente_coso:  Optional[str] = None,
    nivel_riesgo:     Optional[str] = None,
    estado_hallazgo:  Optional[str] = None,
    estado_actividad: Optional[str] = None,
    estado_control:   Optional[str] = None,
    resultado_ev:     Optional[str] = None,
    control_id:       Optional[int] = None,
    fecha_desde:      Optional[str] = None,
    fecha_hasta:      Optional[str] = None,
    q:                Optional[str] = None,
):
    from fastapi_modulo.modulos.control_interno.reporte_store import datos_reporte
    return JSONResponse(datos_reporte(
        tipo=tipo, anio=anio, componente_coso=componente_coso,
        nivel_riesgo=nivel_riesgo, estado_hallazgo=estado_hallazgo,
        estado_actividad=estado_actividad, estado_control=estado_control,
        resultado_ev=resultado_ev, control_id=control_id,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, q=q,
    ))


# ── API: exportar Excel ───────────────────────────────────────────────────────

@router.get("/api/ci-reporte/excel")
def api_excel(
    tipo:             Optional[str] = "completo",
    anio:             Optional[int] = None,
    componente_coso:  Optional[str] = None,
    nivel_riesgo:     Optional[str] = None,
    estado_hallazgo:  Optional[str] = None,
    estado_actividad: Optional[str] = None,
    estado_control:   Optional[str] = None,
    resultado_ev:     Optional[str] = None,
    control_id:       Optional[int] = None,
    fecha_desde:      Optional[str] = None,
    fecha_hasta:      Optional[str] = None,
    q:                Optional[str] = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from fastapi_modulo.modulos.control_interno.reporte_store import datos_reporte

    datos = datos_reporte(
        tipo=tipo, anio=anio, componente_coso=componente_coso,
        nivel_riesgo=nivel_riesgo, estado_hallazgo=estado_hallazgo,
        estado_actividad=estado_actividad, estado_control=estado_control,
        resultado_ev=resultado_ev, control_id=control_id,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, q=q,
    )

    wb = Workbook()
    # Estilos reutilizables
    header_fill   = PatternFill(fill_type="solid", fgColor="1E3A5F")
    header_font   = Font(color="FFFFFF", bold=True, size=10)
    title_font    = Font(bold=True, size=11, color="1E3A5F")
    alt_fill      = PatternFill(fill_type="solid", fgColor="F0F4FA")
    thin_border   = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _write_sheet(ws, titulo: str, filas):
        """Escribe una hoja estilizada."""
        if not filas:
            ws.append(["Sin datos para los filtros aplicados."])
            return
        cols = list(filas[0].keys())
        # Título
        ws.append([titulo])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(1, 1).font      = title_font
        ws.cell(1, 1).alignment = center
        ws.row_dimensions[1].height = 22
        # Generado en
        ws.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
        ws.cell(2, 1).font      = Font(italic=True, size=9, color="888888")
        ws.cell(2, 1).alignment = center
        # Fila vacía
        ws.append([])
        # Encabezados
        ws.append(cols)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(4, col_idx, col_name)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
            cell.border    = thin_border
        ws.row_dimensions[4].height = 18
        # Datos
        for row_idx, row in enumerate(filas, 5):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col in enumerate(cols, 1):
                cell = ws.cell(row_idx, col_idx, row.get(col, ""))
                if fill:
                    cell.fill = fill
                cell.alignment = left
                cell.border    = thin_border
                cell.font      = Font(size=9)
        # Auto-ancho (heurística)
        for col_idx, col in enumerate(cols, 1):
            max_len = max((len(str(row.get(col, ""))) for row in filas), default=0)
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(max_len, len(col)) + 3, 50)

    # Hoja de resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    secciones = {
        "controles":   "Catálogo de controles",
        "actividades": "Programa anual — Actividades",
        "evidencias":  "Evidencias de control",
        "hallazgos":   "Hallazgos",
        "acciones":    "Acciones correctivas",
    }
    ws_resumen.append(["Reporte de Control Interno — COSO"])
    ws_resumen.cell(1, 1).font      = Font(bold=True, size=14, color="1E3A5F")
    ws_resumen.cell(1, 1).alignment = Alignment(horizontal="left")
    ws_resumen.append([f"Tipo: {tipo}   |   Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws_resumen.cell(2, 1).font = Font(italic=True, size=10, color="888888")
    ws_resumen.append([])
    ws_resumen.append(["Sección", "Registros"])
    for c in (1, 2):
        ws_resumen.cell(4, c).fill = header_fill
        ws_resumen.cell(4, c).font = header_font
    for key, label in secciones.items():
        filas = datos.get(key)
        if filas is not None:
            ws_resumen.append([label, len(filas)])
    ws_resumen.column_dimensions["A"].width = 40
    ws_resumen.column_dimensions["B"].width = 15

    # Hojas de datos
    for key, label in secciones.items():
        filas = datos.get(key)
        if filas is None:
            continue
        ws = wb.create_sheet(title=label[:31])
        _write_sheet(ws, label, filas)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"CI_Reporte_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ── API: exportar HTML imprimible ─────────────────────────────────────────────

@router.get("/api/ci-reporte/html", response_class=HTMLResponse)
def api_html(
    tipo:             Optional[str] = "completo",
    anio:             Optional[int] = None,
    componente_coso:  Optional[str] = None,
    nivel_riesgo:     Optional[str] = None,
    estado_hallazgo:  Optional[str] = None,
    estado_actividad: Optional[str] = None,
    estado_control:   Optional[str] = None,
    resultado_ev:     Optional[str] = None,
    control_id:       Optional[int] = None,
    fecha_desde:      Optional[str] = None,
    fecha_hasta:      Optional[str] = None,
    q:                Optional[str] = None,
):
    from html import escape
    from fastapi_modulo.modulos.control_interno.reporte_store import datos_reporte

    datos = datos_reporte(
        tipo=tipo, anio=anio, componente_coso=componente_coso,
        nivel_riesgo=nivel_riesgo, estado_hallazgo=estado_hallazgo,
        estado_actividad=estado_actividad, estado_control=estado_control,
        resultado_ev=resultado_ev, control_id=control_id,
        fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, q=q,
    )

    secciones_labels = {
        "controles":   "Catálogo de controles",
        "actividades": "Programa anual — Actividades",
        "evidencias":  "Evidencias de control",
        "hallazgos":   "Hallazgos",
        "acciones":    "Acciones correctivas",
    }

    tablas_html = ""
    for key, label in secciones_labels.items():
        filas = datos.get(key)
        if not filas:
            continue
        cols = list(filas[0].keys())
        header_cells = "".join(f"<th>{escape(str(c))}</th>" for c in cols)
        rows_html = ""
        for i, row in enumerate(filas):
            clase = "alt" if i % 2 else ""
            cells = "".join(f"<td>{escape(str(row.get(c, '')))}</td>" for c in cols)
            rows_html += f"<tr class='{clase}'>{cells}</tr>"
        tablas_html += f"""
        <section class="seccion">
          <h2>{escape(label)} <span class="count">({len(filas)} registros)</span></h2>
          <div class="table-wrap">
            <table>
              <thead><tr>{header_cells}</tr></thead>
              <tbody>{rows_html}</tbody>
            </table>
          </div>
        </section>"""

    titulo_tipo = {
        "completo":   "Informe completo",
        "controles":  "Catálogo de controles",
        "programa":   "Programa anual",
        "evidencias": "Evidencias",
        "hallazgos":  "Hallazgos y acciones",
    }.get(tipo or "completo", tipo or "completo")

    filtros_activos = []
    if anio:            filtros_activos.append(f"Año: {anio}")
    if componente_coso: filtros_activos.append(f"Componente COSO: {componente_coso}")
    if nivel_riesgo:    filtros_activos.append(f"Nivel de riesgo: {nivel_riesgo}")
    if estado_hallazgo: filtros_activos.append(f"Estado hallazgo: {estado_hallazgo}")
    if resultado_ev:    filtros_activos.append(f"Resultado evidencia: {resultado_ev}")
    if fecha_desde:     filtros_activos.append(f"Desde: {fecha_desde}")
    if fecha_hasta:     filtros_activos.append(f"Hasta: {fecha_hasta}")
    if q:               filtros_activos.append(f"Búsqueda: {q}")
    filtros_str = " · ".join(filtros_activos) or "Sin filtros aplicados"

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Reporte — {escape(titulo_tipo)}</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; }}
    body {{ font-family: Arial, Helvetica, sans-serif; font-size: 11px; color: #222; margin: 0; padding: 20px 30px; }}
    .header {{ border-bottom: 3px solid #1E3A5F; padding-bottom: 12px; margin-bottom: 18px; }}
    .header h1 {{ color: #1E3A5F; font-size: 20px; margin: 0 0 4px; }}
    .header .meta {{ color: #777; font-size: 10px; }}
    .filtros {{ background: #f5f7fa; border: 1px solid #dde3ee; border-radius: 4px; padding: 8px 12px; margin-bottom: 20px; font-size: 10px; color: #555; }}
    .seccion {{ margin-bottom: 24px; page-break-inside: avoid; }}
    .seccion h2 {{ background: #1E3A5F; color: white; padding: 7px 12px; font-size: 12px; margin: 0 0 6px; border-radius: 3px 3px 0 0; }}
    .seccion h2 .count {{ font-weight: normal; opacity: .8; font-size: 10px; }}
    .table-wrap {{ overflow-x: auto; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
    thead tr {{ background: #2d5494; color: white; }}
    thead th {{ padding: 5px 8px; text-align: left; font-weight: bold; border: 1px solid #244280; white-space: nowrap; }}
    tbody td {{ padding: 4px 8px; border: 1px solid #ddd; vertical-align: top; }}
    tbody tr.alt {{ background: #f0f4fa; }}
    .footer {{ border-top: 1px solid #ddd; padding-top: 8px; margin-top: 20px; color: #aaa; font-size: 9px; text-align: right; }}
    @media print {{
      body {{ padding: 10px; }}
      .no-print {{ display: none !important; }}
      @page {{ margin: 1.5cm; }}
    }}
  </style>
</head>
<body>
  <div class="header">
    <h1>Sistema de Control Interno COSO — {escape(titulo_tipo)}</h1>
    <div class="meta">Generado el {escape(datetime.now().strftime('%d/%m/%Y a las %H:%M'))} hrs</div>
  </div>
  <div class="filtros">
    <strong>Filtros aplicados:</strong> {escape(filtros_str)}
  </div>
  <div class="no-print" style="margin-bottom:16px">
    <button onclick="window.print()" style="padding:8px 16px;background:#1E3A5F;color:white;border:none;border-radius:4px;cursor:pointer;font-size:12px;">
      🖨️ Imprimir
    </button>
  </div>
  {tablas_html if tablas_html else '<p style="color:#999;text-align:center;padding:40px">No hay datos para los filtros seleccionados.</p>'}
  <div class="footer">SIPET — Reporte de Control Interno · {datetime.now().year}</div>
</body>
</html>"""

    return HTMLResponse(content=html)
